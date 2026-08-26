"""Sheets Agent - creates Google Sheets and Microsoft Excel (.xlsx) spreadsheets."""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any, Dict, List, Optional
from uuid import uuid4

import structlog

from ..db.database import Database
from ..llm.base import BaseLLM, Message
from ..workspace.manager import WorkspaceManager
from .base import BaseAgent

logger = structlog.get_logger()


class SheetsAgent(BaseAgent):
    """Creates Google Sheets for campaign trackers, budgets, and analytics.

    Uses the Google Sheets API (via service account) to create live, editable
    spreadsheets and shares them via link.  Falls back gracefully when
    credentials are not configured.
    """

    agent_id = "sheets_agent"
    agent_name = "Sheets Agent"

    def __init__(
        self,
        llm: BaseLLM,
        db: Database,
        workspace_manager: WorkspaceManager,
        google_credentials_path: str = "",
        google_oauth_token_path: str = "",
        media_storage: Optional[Any] = None,
        proofreader: Optional[Any] = None,
        max_iterations: int = 10,
    ) -> None:
        self._workspace_mgr = workspace_manager
        self._google_credentials_path = google_credentials_path
        self._google_oauth_token_path = google_oauth_token_path
        self._media_storage = media_storage
        self._proofreader = proofreader
        self._sheets_service: Optional[Any] = None
        self._drive_service: Optional[Any] = None
        super().__init__(llm=llm, db=db, max_iterations=max_iterations)

    # ── Google API lazy init ──────────────────────────────────────────────

    def _get_google_services(self) -> bool:
        """Lazily initialize Google Sheets and Drive API services.

        Returns True if services are available, False otherwise.
        Uses OAuth2 token (preferred) or service account credentials (fallback).
        """
        if self._sheets_service is not None:
            return True

        try:
            from googleapiclient.discovery import build

            from ..google_auth import get_google_credentials

            SCOPES = [
                "https://www.googleapis.com/auth/spreadsheets",
                "https://www.googleapis.com/auth/drive.file",
            ]
            credentials = get_google_credentials(
                oauth_token_path=self._google_oauth_token_path,
                service_account_path=self._google_credentials_path,
                scopes=SCOPES,
            )
            if credentials is None:
                return False

            self._sheets_service = build("sheets", "v4", credentials=credentials)
            self._drive_service = build("drive", "v3", credentials=credentials)
            logger.info("sheets_google_services_initialized")
            return True
        except Exception as e:
            logger.error("sheets_google_services_init_failed", error=str(e))
            return False

    def _share_spreadsheet(self, spreadsheet_id: str) -> None:
        """Share a spreadsheet with authorized users only (restricted access)."""
        from ..google_auth import share_file_restricted

        share_file_restricted(self._drive_service, spreadsheet_id)

    # ── Output format dispatch ────────────────────────────────────────────

    def _resolve_format(self, requested: str) -> str:
        """Determine output format. Default to google_sheets when credentials exist.

        Accepted values:
            ""              auto: google_sheets if creds, otherwise xlsx
            "google_sheets" Google Sheets only
            "xlsx"          Microsoft Excel only
            "both"          Build both formats
        """
        if requested:
            r = requested.lower().strip()
            if r in ("google", "gsheet", "gsheets"):
                return "google_sheets"
            if r in ("excel", "office", "ms_excel", "microsoft_excel"):
                return "xlsx"
            return r
        if self._google_credentials_path and Path(self._google_credentials_path).exists():
            return "google_sheets"
        return "xlsx"

    async def _build_and_export(
        self,
        structure: Dict[str, Any],
        workspace_id: str,
        output_format: str,
    ) -> Dict[str, Any]:
        """Build the spreadsheet and export in the requested format(s)."""
        if output_format == "google_sheets":
            return self._build_google_sheet(structure, workspace_id)

        if output_format == "xlsx":
            return await self._build_xlsx(structure, workspace_id)

        if output_format == "both":
            gsheet_result = self._build_google_sheet(structure, workspace_id)
            xlsx_result = await self._build_xlsx(structure, workspace_id)
            merged: Dict[str, Any] = {"status": "created"}
            if gsheet_result.get("google_sheets_url"):
                merged["google_sheets_url"] = gsheet_result["google_sheets_url"]
                merged["google_sheets_id"] = gsheet_result.get("google_sheets_id")
            if xlsx_result.get("file_path"):
                merged["file_path"] = xlsx_result["file_path"]
            if xlsx_result.get("cdn_url"):
                merged["cdn_url"] = xlsx_result["cdn_url"]
            if (
                gsheet_result.get("status") != "created"
                and xlsx_result.get("status") != "created"
            ):
                merged["status"] = "error"
                merged["message"] = "Both Google Sheets and .xlsx generation failed"
            return merged

        return {
            "status": "error",
            "message": f"Unknown output_format: {output_format!r}",
        }

    # ── System prompt ─────────────────────────────────────────────────────

    def get_system_prompt(self, workspace_id: Optional[str] = None) -> str:
        ws_label = workspace_id or "all workspaces"
        gsheets_available = bool(self._google_credentials_path)
        return f"""You are the Sheets Agent for {ws_label}.
You create professional Google Sheets for marketing teams.

Google Sheets API: {"configured" if gsheets_available else "not configured"}

CRITICAL RULES:
- You MUST ALWAYS call the create_spreadsheet, create_campaign_tracker, or create_budget_sheet tool. NEVER generate spreadsheet content as text in your response.
- Your ONLY job is to call a tool that creates a real Google Sheet and return the link. Do NOT write out table data yourself.
- If the tool returns "not_configured", report that Google credentials need to be set up. Do NOT fall back to generating content as text.
- After creating a spreadsheet, your response should be brief: state that the spreadsheet was created and include the Google Sheets URL.
- ANTI-AI WRITING (MANDATORY): NEVER use em dashes (—) or en dashes (–). NEVER bold words mid-sentence. Use commas, periods, semicolons, or " - " instead. Bolding is only for headers or standalone labels.

You can create:
- Campaign performance trackers
- Budget allocation and spending sheets
- Content calendars
- Analytics dashboards
- ROI calculation sheets
- A/B test result trackers
- Lead scoring matrices

When creating spreadsheets:
- Use clear column headers
- Include formulas for totals, averages, and percentages where appropriate
- Format numbers, currencies, and dates properly
- Use bold headers and alternating row colors for readability
- Add summary rows with aggregation formulas

Output is a live Google Sheets link that opens directly in the browser.
Current workspace: {ws_label}"""

    # ── Tools ─────────────────────────────────────────────────────────────

    def register_tools(self) -> None:

        @self.tool_registry.register(
            name="read_spreadsheet",
            description=(
                "Read data from an existing XLSX or CSV file. Returns structured headers "
                "and rows. Use this to analyze existing spreadsheets, import data, or "
                "reference data when creating new sheets. "
                "Relative paths are resolved against the data/ directory."
            ),
        )
        async def read_spreadsheet(
            file_path: str,
            sheet_name: Optional[str] = None,
            max_rows: int = 500,
        ) -> Dict[str, Any]:
            from ..files.spreadsheet import read_spreadsheet as _read

            return _read(file_path, sheet_name=sheet_name, max_rows=max_rows)

        @self.tool_registry.register(
            name="list_data_files",
            description=(
                "List available spreadsheet and data files in the data directory. "
                "Use this to discover what files are available before reading them."
            ),
        )
        async def list_data_files(
            directory: Optional[str] = None,
        ) -> Dict[str, Any]:
            from ..files.spreadsheet import list_data_files as _list

            return _list(directory=directory)

        @self.tool_registry.register(
            name="create_spreadsheet",
            description=(
                "Create a spreadsheet from a topic. Generates a well-structured spreadsheet "
                "with headers, sample data, and formulas. "
                "Defaults to Google Sheets when configured, otherwise .xlsx. "
                "Set output_format to 'google_sheets', 'xlsx', or 'both'."
            ),
        )
        async def create_spreadsheet(
            topic: str,
            workspace_id: str,
            purpose: str = "tracking",
            columns: Optional[List[str]] = None,
            rows: Optional[int] = None,
            output_format: str = "",
        ) -> Dict[str, Any]:
            fmt = self._resolve_format(output_format)
            brand_voice = await self._workspace_mgr.get_brand_voice(workspace_id)

            structure = await self._generate_sheet_structure(
                topic=topic,
                purpose=purpose,
                brand_voice=brand_voice,
                columns=columns,
                rows=rows,
            )

            if self._proofreader:
                try:
                    structure, _ = await self._proofreader.proofread_dict_fields(
                        structure,
                        fields=["title", "name", "headers"],
                    )
                except Exception as e:
                    logger.warning("sheets_proofread_failed", error=str(e))

            result = await self._build_and_export(structure, workspace_id, fmt)
            return {
                **result,
                "topic": topic,
                "purpose": purpose,
                "workspace_id": workspace_id,
                "output_format": fmt,
            }

        @self.tool_registry.register(
            name="create_campaign_tracker",
            description=(
                "Create a pre-structured campaign tracking spreadsheet with platform columns, "
                "metrics, and summary formulas. "
                "Defaults to Google Sheets when configured, otherwise .xlsx. "
                "Set output_format to 'google_sheets', 'xlsx', or 'both'."
            ),
        )
        async def create_campaign_tracker(
            workspace_id: str,
            campaign_name: str,
            platforms: List[str],
            metrics: List[str],
            period: str = "monthly",
            output_format: str = "",
        ) -> Dict[str, Any]:
            fmt = self._resolve_format(output_format)
            prompt = f"""Create a campaign tracking spreadsheet structure.

CAMPAIGN: {campaign_name}
PLATFORMS: {", ".join(platforms)}
METRICS: {", ".join(metrics)}
PERIOD: {period}

Create a sheet with:
- Headers row with: Date, Platform, {", ".join(metrics)}
- Sample rows for each platform (use realistic placeholder data)
- A summary row at the bottom with SUM/AVERAGE formulas
- Formulas should reference the actual cell ranges

{_sheet_structure_format_instructions()}"""

            response = await self.llm.complete(
                messages=[Message(role="user", content=prompt)],
                temperature=0.5,
            )

            structure = _parse_sheet_structure(response.content)
            if not structure.get("title"):
                structure["title"] = f"{campaign_name} - Campaign Tracker"

            if self._proofreader:
                try:
                    structure, _ = await self._proofreader.proofread_dict_fields(
                        structure,
                        fields=["title", "name", "headers"],
                    )
                except Exception as e:
                    logger.warning("sheets_proofread_failed", error=str(e))

            result = await self._build_and_export(structure, workspace_id, fmt)
            return {
                **result,
                "campaign_name": campaign_name,
                "platforms": platforms,
                "workspace_id": workspace_id,
                "output_format": fmt,
            }

        @self.tool_registry.register(
            name="create_budget_sheet",
            description=(
                "Create a budget allocation spreadsheet with categories, subtotals, and "
                "percentage formulas. "
                "Defaults to Google Sheets when configured, otherwise .xlsx. "
                "Set output_format to 'google_sheets', 'xlsx', or 'both'."
            ),
        )
        async def create_budget_sheet(
            workspace_id: str,
            budget_total: float,
            categories: List[str],
            period: str = "quarterly",
            output_format: str = "",
        ) -> Dict[str, Any]:
            fmt = self._resolve_format(output_format)
            prompt = f"""Create a marketing budget spreadsheet structure.

TOTAL BUDGET: ${budget_total:,.2f}
CATEGORIES: {", ".join(categories)}
PERIOD: {period}

Create a sheet with:
- Headers: Category, Allocated Budget, Spent, Remaining, % of Total, % Spent
- One row per category with realistic allocation amounts that sum to the total budget
- A TOTAL row at the bottom with SUM formulas
- Formulas for Remaining (Allocated - Spent), % of Total, and % Spent
- Reference actual cell ranges in formulas (e.g., =B2-C2, =B2/B${{last_row}})

{_sheet_structure_format_instructions()}"""

            response = await self.llm.complete(
                messages=[Message(role="user", content=prompt)],
                temperature=0.5,
            )

            structure = _parse_sheet_structure(response.content)
            if not structure.get("title"):
                structure["title"] = f"Marketing Budget - {period.title()}"

            if self._proofreader:
                try:
                    structure, _ = await self._proofreader.proofread_dict_fields(
                        structure,
                        fields=["title", "name", "headers"],
                    )
                except Exception as e:
                    logger.warning("sheets_proofread_failed", error=str(e))

            result = await self._build_and_export(structure, workspace_id, fmt)
            return {
                **result,
                "budget_total": budget_total,
                "categories": categories,
                "workspace_id": workspace_id,
                "output_format": fmt,
            }

    # ── Internal methods ──────────────────────────────────────────────────

    async def _generate_sheet_structure(
        self,
        topic: str,
        purpose: str,
        brand_voice: str,
        columns: Optional[List[str]] = None,
        rows: Optional[int] = None,
    ) -> Dict[str, Any]:
        """Use the LLM to generate a structured JSON spreadsheet outline."""
        columns_hint = ""
        if columns:
            columns_hint = f"\nREQUESTED COLUMNS: {', '.join(columns)}"
        rows_hint = ""
        if rows:
            rows_hint = f"\nNUMBER OF DATA ROWS: {rows}"

        prompt = f"""Create a professional spreadsheet structure.

TOPIC: {topic}
PURPOSE: {purpose}{columns_hint}{rows_hint}

BRAND CONTEXT:
{brand_voice[:300] if brand_voice else "No brand voice specified."}

{_sheet_structure_format_instructions()}"""

        response = await self.llm.complete(
            messages=[Message(role="user", content=prompt)],
            temperature=0.5,
        )

        return _parse_sheet_structure(response.content)

    def _build_google_sheet(self, structure: Dict[str, Any], workspace_id: str) -> Dict[str, Any]:
        """Create a Google Sheet via the Sheets API and return its URL."""
        if not self._get_google_services():
            return {
                "status": "not_configured",
                "message": (
                    "Google credentials not configured. "
                    "Set GOOGLE_CREDENTIALS_PATH in .env to a service account JSON file."
                ),
            }

        title = structure.get("title", "Untitled Spreadsheet")
        sheets_data = structure.get("sheets", [])
        if not sheets_data:
            sheets_data = [
                {
                    "name": "Sheet1",
                    "headers": ["Column A", "Column B", "Column C"],
                    "rows": [],
                }
            ]

        try:
            # 1. Build sheet definitions for creation
            sheet_defs = []
            for i, sd in enumerate(sheets_data):
                sheet_defs.append(
                    {
                        "properties": {
                            "title": sd.get("name", f"Sheet{i + 1}"),
                            "index": i,
                        }
                    }
                )

            spreadsheet = (
                self._sheets_service.spreadsheets()
                .create(
                    body={
                        "properties": {"title": title},
                        "sheets": sheet_defs,
                    }
                )
                .execute()
            )
            spreadsheet_id = spreadsheet["spreadsheetId"]

            # 2. Populate data for each sheet
            value_ranges = []
            for i, sd in enumerate(sheets_data):
                sheet_name = sd.get("name", f"Sheet{i + 1}")
                headers = sd.get("headers", [])
                rows = sd.get("rows", [])
                formulas = sd.get("formulas", {})

                all_rows: List[List[Any]] = []
                if headers:
                    all_rows.append(headers)
                all_rows.extend(rows)

                # Apply formulas to specific cells
                if formulas:
                    for cell_ref, formula in formulas.items():
                        row_idx, col_idx = _parse_cell_ref(cell_ref)
                        if row_idx is not None:
                            # Extend rows if needed
                            while len(all_rows) <= row_idx:
                                all_rows.append([])
                            while len(all_rows[row_idx]) <= col_idx:
                                all_rows[row_idx].append("")
                            all_rows[row_idx][col_idx] = formula

                if all_rows:
                    value_ranges.append(
                        {
                            "range": f"'{sheet_name}'!A1",
                            "values": all_rows,
                        }
                    )

            if value_ranges:
                self._sheets_service.spreadsheets().values().batchUpdate(
                    spreadsheetId=spreadsheet_id,
                    body={
                        "valueInputOption": "USER_ENTERED",
                        "data": value_ranges,
                    },
                ).execute()

            # 3. Apply formatting (bold headers, column widths)
            format_requests = self._build_format_requests(spreadsheet, sheets_data)
            if format_requests:
                self._sheets_service.spreadsheets().batchUpdate(
                    spreadsheetId=spreadsheet_id,
                    body={"requests": format_requests},
                ).execute()

            # 4. Share spreadsheet
            self._share_spreadsheet(spreadsheet_id)

            # 5. Move into CMO Agent folder
            from ..google_auth import ensure_drive_folder, move_file_to_folder

            folder_id = ensure_drive_folder(self._drive_service)
            if folder_id:
                move_file_to_folder(self._drive_service, spreadsheet_id, folder_id)

            sheets_url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit"
            logger.info(
                "google_sheet_created",
                spreadsheet_id=spreadsheet_id,
                sheets=len(sheets_data),
                workspace=workspace_id,
            )

            return {
                "status": "created",
                "google_sheets_url": sheets_url,
                "google_sheets_id": spreadsheet_id,
            }

        except Exception as e:
            logger.error("google_sheet_build_failed", error=str(e))
            return {"status": "error", "message": str(e)}

    async def _build_xlsx(
        self, structure: Dict[str, Any], workspace_id: str
    ) -> Dict[str, Any]:
        """Build a Microsoft Excel (.xlsx) file mirroring the sheet structure.

        Mirrors the same structure consumed by _build_google_sheet:
            { title, sheets: [{ name, headers, column_types, rows,
                                column_widths, formulas }] }
        Supports column_types: text, currency, percentage, number, integer, date.
        Saves to data/sheets/ and optionally uploads to the configured media store.
        """
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter

        _NUMBER_FORMATS = {
            "currency": "$#,##0.00",
            "percentage": "0.0%",
            "number": "#,##0.00",
            "integer": "#,##0",
            "date": "yyyy-mm-dd",
            "text": "@",
        }

        title = structure.get("title", "Untitled Spreadsheet")
        sheets_data = structure.get("sheets", [])
        if not sheets_data:
            sheets_data = [{"name": "Sheet1", "headers": [], "rows": []}]

        try:
            wb = Workbook()
            # Remove the default sheet created by Workbook()
            wb.remove(wb.active)

            header_font = Font(bold=True, color="FFFFFFFF")
            header_fill = PatternFill(
                start_color="FF002F6C", end_color="FF002F6C", fill_type="solid"
            )
            thin_border = Border(
                left=Side(style="thin", color="FFD0D0D0"),
                right=Side(style="thin", color="FFD0D0D0"),
                top=Side(style="thin", color="FFD0D0D0"),
                bottom=Side(style="thin", color="FFD0D0D0"),
            )

            for sheet_def in sheets_data:
                name = sheet_def.get("name", "Sheet1")[:31]  # Excel limit
                headers = sheet_def.get("headers", [])
                rows = sheet_def.get("rows", [])
                column_types = sheet_def.get("column_types", [])
                column_widths = sheet_def.get("column_widths", [])
                formulas = sheet_def.get("formulas", {})

                ws = wb.create_sheet(name)

                # Header row
                for ci, header in enumerate(headers, start=1):
                    cell = ws.cell(row=1, column=ci, value=str(header))
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                    cell.border = thin_border

                # Data rows
                for ri, row_data in enumerate(rows, start=2):
                    for ci, value in enumerate(row_data, start=1):
                        cell = ws.cell(row=ri, column=ci, value=value)
                        cell.border = thin_border
                        # Apply column type number format
                        col_idx = ci - 1
                        if col_idx < len(column_types):
                            col_type = str(column_types[col_idx]).lower()
                            if col_type in _NUMBER_FORMATS:
                                cell.number_format = _NUMBER_FORMATS[col_type]

                # Formulas (e.g., {"C10": "=SUM(C2:C9)"})
                for cell_ref, formula in formulas.items():
                    try:
                        cell = ws[cell_ref]
                        # Make sure the formula starts with =
                        f_str = formula if formula.startswith("=") else f"={formula}"
                        cell.value = f_str
                        cell.font = Font(bold=True)
                        cell.border = thin_border
                    except (ValueError, KeyError):
                        logger.warning("xlsx_formula_skipped", cell=cell_ref, formula=formula)

                # Column widths
                for ci, width in enumerate(column_widths, start=1):
                    try:
                        # Excel widths are roughly pixels/7; Google Sheets uses pixels.
                        ws.column_dimensions[get_column_letter(ci)].width = (
                            float(width) / 7 if width else 12
                        )
                    except (ValueError, TypeError):
                        pass

                # Default reasonable widths if not specified
                if not column_widths and headers:
                    for ci in range(1, len(headers) + 1):
                        ws.column_dimensions[get_column_letter(ci)].width = 18

                # Freeze the header row
                ws.freeze_panes = "A2"

            sheets_dir = Path("./data/sheets")
            sheets_dir.mkdir(parents=True, exist_ok=True)
            slug = re.sub(r"[^a-z0-9]+", "-", title.lower()).strip("-")[:40] or "spreadsheet"
            filename = f"{workspace_id}_{slug}_{uuid4().hex[:6]}.xlsx"
            file_path = sheets_dir / filename
            wb.save(str(file_path))

            logger.info(
                "xlsx_created",
                file_path=str(file_path),
                sheets=len(sheets_data),
                workspace=workspace_id,
            )

            result: Dict[str, Any] = {
                "status": "created",
                "file_path": str(file_path),
            }

            if self._media_storage:
                try:
                    asset = await self._media_storage.upload_from_file(
                        file_path=str(file_path),
                        workspace_id=workspace_id,
                        media_type="sheet",
                        source_agent=self.agent_id,
                        prompt=title,
                    )
                    if asset and asset.cdn_url:
                        result["cdn_url"] = asset.cdn_url
                except Exception as e:
                    logger.warning("xlsx_cdn_upload_failed", error=str(e))

            return result

        except Exception as e:
            logger.error("xlsx_build_failed", error=str(e))
            return {"status": "error", "message": str(e)}

    @staticmethod
    def _build_format_requests(
        spreadsheet: Dict[str, Any], sheets_data: List[Dict[str, Any]]
    ) -> List[Dict[str, Any]]:
        """Build formatting requests: bold headers, borders, banding, number formats."""
        requests: List[Dict[str, Any]] = []
        created_sheets = spreadsheet.get("sheets", [])

        for i, sd in enumerate(sheets_data):
            if i >= len(created_sheets):
                break
            sheet_id = created_sheets[i]["properties"]["sheetId"]
            headers = sd.get("headers", [])
            rows = sd.get("rows", [])
            column_widths = sd.get("column_widths", [])
            column_types = sd.get("column_types", [])
            num_cols = len(headers) if headers else 0
            # +1 for header row
            num_rows = (1 + len(rows)) if headers else len(rows)

            # Bold header row with background
            if headers:
                requests.append(
                    {
                        "repeatCell": {
                            "range": {
                                "sheetId": sheet_id,
                                "startRowIndex": 0,
                                "endRowIndex": 1,
                                "startColumnIndex": 0,
                                "endColumnIndex": num_cols,
                            },
                            "cell": {
                                "userEnteredFormat": {
                                    "textFormat": {"bold": True},
                                    "backgroundColor": {
                                        "red": 0.9,
                                        "green": 0.9,
                                        "blue": 0.9,
                                    },
                                    "horizontalAlignment": "CENTER",
                                }
                            },
                            "fields": "userEnteredFormat(textFormat,backgroundColor,"
                            "horizontalAlignment)",
                        }
                    }
                )

                # Freeze header row
                requests.append(
                    {
                        "updateSheetProperties": {
                            "properties": {
                                "sheetId": sheet_id,
                                "gridProperties": {"frozenRowCount": 1},
                            },
                            "fields": "gridProperties.frozenRowCount",
                        }
                    }
                )

            # Borders — light gray gridlines around entire data range
            if num_cols > 0 and num_rows > 0:
                border_style = {
                    "style": "SOLID",
                    "width": 1,
                    "color": {"red": 0.8, "green": 0.8, "blue": 0.8},
                }
                data_range = {
                    "sheetId": sheet_id,
                    "startRowIndex": 0,
                    "endRowIndex": num_rows,
                    "startColumnIndex": 0,
                    "endColumnIndex": num_cols,
                }
                requests.append(
                    {
                        "updateBorders": {
                            "range": data_range,
                            "top": border_style,
                            "bottom": border_style,
                            "left": border_style,
                            "right": border_style,
                            "innerHorizontal": border_style,
                            "innerVertical": border_style,
                        }
                    }
                )

            # Alternating row colors (banding)
            if num_cols > 0 and num_rows > 1:
                requests.append(
                    {
                        "addBanding": {
                            "bandedRange": {
                                "range": {
                                    "sheetId": sheet_id,
                                    "startRowIndex": 0,
                                    "endRowIndex": num_rows,
                                    "startColumnIndex": 0,
                                    "endColumnIndex": num_cols,
                                },
                                "rowProperties": {
                                    "headerColor": {
                                        "red": 0.9,
                                        "green": 0.9,
                                        "blue": 0.9,
                                    },
                                    "firstBandColor": {
                                        "red": 1.0,
                                        "green": 1.0,
                                        "blue": 1.0,
                                    },
                                    "secondBandColor": {
                                        "red": 0.94,
                                        "green": 0.96,
                                        "blue": 1.0,
                                    },
                                },
                            }
                        }
                    }
                )

            # Number formatting and alignment per column_types
            number_formats: Dict[str, tuple] = {
                "currency": ("$#,##0.00", "RIGHT"),
                "percentage": ("0.0%", "RIGHT"),
                "number": ("#,##0.00", "RIGHT"),
                "integer": ("#,##0", "RIGHT"),
                "date": ("yyyy-mm-dd", "CENTER"),
                "text": ("", "LEFT"),
            }
            for j, col_type in enumerate(column_types):
                if j >= num_cols:
                    break
                fmt_info = number_formats.get(col_type)
                if not fmt_info:
                    continue
                pattern, alignment = fmt_info
                cell_format: Dict[str, Any] = {"horizontalAlignment": alignment}
                fields = "userEnteredFormat.horizontalAlignment"
                if pattern:
                    cell_format["numberFormat"] = {"type": "NUMBER", "pattern": pattern}
                    fields += ",userEnteredFormat.numberFormat"
                # Apply to data rows only (skip header)
                if num_rows > 1:
                    requests.append(
                        {
                            "repeatCell": {
                                "range": {
                                    "sheetId": sheet_id,
                                    "startRowIndex": 1,
                                    "endRowIndex": num_rows,
                                    "startColumnIndex": j,
                                    "endColumnIndex": j + 1,
                                },
                                "cell": {"userEnteredFormat": cell_format},
                                "fields": fields,
                            }
                        }
                    )

            # Column widths
            for j, width in enumerate(column_widths):
                if isinstance(width, (int, float)) and width > 0:
                    requests.append(
                        {
                            "updateDimensionProperties": {
                                "range": {
                                    "sheetId": sheet_id,
                                    "dimension": "COLUMNS",
                                    "startIndex": j,
                                    "endIndex": j + 1,
                                },
                                "properties": {"pixelSize": int(width)},
                                "fields": "pixelSize",
                            }
                        }
                    )

            # Auto-resize columns when no explicit widths provided
            if not column_widths and num_cols > 0:
                requests.append(
                    {
                        "autoResizeDimensions": {
                            "dimensions": {
                                "sheetId": sheet_id,
                                "dimension": "COLUMNS",
                                "startIndex": 0,
                                "endIndex": num_cols,
                            }
                        }
                    }
                )

        return requests


# ══════════════════════════════════════════════════════════════════════════
# Helpers
# ══════════════════════════════════════════════════════════════════════════


def _parse_cell_ref(ref: str) -> tuple:
    """Parse a cell reference like 'C10' into (row_index, col_index) zero-based."""
    match = re.match(r"^([A-Z]+)(\d+)$", ref.upper())
    if not match:
        return (None, None)
    col_letters = match.group(1)
    row_num = int(match.group(2))
    col_idx = 0
    for ch in col_letters:
        col_idx = col_idx * 26 + (ord(ch) - ord("A"))
    return (row_num - 1, col_idx)


def _sheet_structure_format_instructions() -> str:
    return """Return ONLY a valid JSON object with this structure:
{
  "title": "Spreadsheet Title",
  "sheets": [
    {
      "name": "Sheet1",
      "headers": ["Column A", "Column B", "Column C"],
      "column_types": ["text", "currency", "percentage"],
      "rows": [
        ["value1", 1500.00, 0.25],
        ["value2", 3200.00, 0.45]
      ],
      "column_widths": [150, 120, 100],
      "formulas": {
        "C10": "=SUM(C2:C9)",
        "D10": "=AVERAGE(D2:D9)"
      }
    }
  ]
}

column_types tells the formatter how to format each column. Supported types:
  "text" (default), "currency" ($#,##0.00), "percentage" (0.0%),
  "number" (#,##0.00), "integer" (#,##0), "date" (yyyy-mm-dd).

Use realistic sample data. Include formulas for totals and averages where appropriate.
Reference actual cell ranges in formulas."""


def _parse_sheet_structure(raw: str) -> Dict[str, Any]:
    """Parse LLM output into a spreadsheet structure dict."""
    try:
        return json.loads(raw)
    except (json.JSONDecodeError, TypeError):
        pass

    match = re.search(r"```(?:json)?\s*\n?(.*?)\n?```", raw, re.DOTALL)
    if match:
        try:
            return json.loads(match.group(1))
        except (json.JSONDecodeError, TypeError):
            pass

    logger.warning("sheet_structure_parse_fallback", raw_length=len(raw))
    return {
        "title": "Untitled Spreadsheet",
        "sheets": [
            {
                "name": "Sheet1",
                "headers": ["Item", "Value", "Notes"],
                "rows": [["(data to be added)", "", ""]],
            }
        ],
    }
