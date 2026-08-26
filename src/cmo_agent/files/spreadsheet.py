"""Spreadsheet reading utilities for agent tools.

Provides structured reading of XLSX and CSV files so agents can access
spreadsheet data without manual conversion to text files.
"""

from __future__ import annotations

import csv
import io
import os
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import structlog

logger = structlog.get_logger()

# Limit rows to stay within LLM context
MAX_ROWS = 500

# Base data directory (project root / data)
_DATA_DIR = Path(__file__).resolve().parents[3] / "data"


def read_spreadsheet(
    file_path: str,
    sheet_name: Optional[str] = None,
    max_rows: int = MAX_ROWS,
) -> Dict[str, Any]:
    """Read an XLSX or CSV file and return structured data.

    Args:
        file_path: Absolute or relative path to the file. Relative paths
                   are resolved against the project ``data/`` directory.
        sheet_name: For XLSX files, the sheet to read. Defaults to first sheet.
        max_rows: Maximum data rows to return (default 500).

    Returns:
        Dict with keys: file, sheet, headers, rows, row_count, truncated,
        and available_sheets (XLSX only).
    """
    path = _resolve_path(file_path)
    if not path.exists():
        return {"error": f"File not found: {file_path}"}

    ext = path.suffix.lower()
    if ext == ".csv":
        return _read_csv(path, max_rows)
    if ext in (".xlsx", ".xls"):
        return _read_xlsx(path, sheet_name, max_rows)
    return {"error": f"Unsupported file type: {ext}. Supported: .xlsx, .csv"}


def list_data_files(
    directory: Optional[str] = None,
    extensions: Optional[List[str]] = None,
) -> Dict[str, Any]:
    """List spreadsheet/data files in a directory.

    Args:
        directory: Directory to scan. Defaults to the project ``data/`` directory.
                   Relative paths are resolved against ``data/``.
        extensions: File extensions to include (default: .xlsx, .csv).

    Returns:
        Dict with files list and count.
    """
    if extensions is None:
        extensions = [".xlsx", ".csv"]
    ext_set = {e.lower() if e.startswith(".") else f".{e.lower()}" for e in extensions}

    if directory:
        scan_dir = _resolve_path(directory)
    else:
        scan_dir = _DATA_DIR

    if not scan_dir.exists() or not scan_dir.is_dir():
        return {"error": f"Directory not found: {directory or 'data/'}", "files": []}

    files: List[Dict[str, Any]] = []
    for root, _dirs, filenames in os.walk(scan_dir):
        for fname in sorted(filenames):
            fpath = Path(root) / fname
            if fpath.suffix.lower() in ext_set:
                stat = fpath.stat()
                files.append(
                    {
                        "name": fname,
                        "path": str(fpath),
                        "relative_path": str(fpath.relative_to(_DATA_DIR))
                        if str(fpath).startswith(str(_DATA_DIR))
                        else str(fpath),
                        "size_kb": round(stat.st_size / 1024, 1),
                        "modified": datetime.fromtimestamp(stat.st_mtime).isoformat(
                            timespec="seconds"
                        ),
                    }
                )

    return {"directory": str(scan_dir), "count": len(files), "files": files}


def list_sheets(file_path: str) -> Dict[str, Any]:
    """List sheet names in an XLSX file.

    Args:
        file_path: Path to the XLSX file.

    Returns:
        Dict with sheets list.
    """
    path = _resolve_path(file_path)
    if not path.exists():
        return {"error": f"File not found: {file_path}"}
    if path.suffix.lower() not in (".xlsx", ".xls"):
        return {"error": "Only XLSX files have multiple sheets"}

    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        names = wb.sheetnames
        wb.close()
        return {"file": str(path), "sheets": names}
    except Exception as e:
        logger.warning("list_sheets_failed", error=str(e))
        return {"error": f"Failed to read sheets: {e}"}


# ── Internal helpers ─────────────────────────────────────────────────────


def _resolve_path(file_path: str) -> Path:
    """Resolve a file path, treating relative paths as relative to data/."""
    p = Path(file_path)
    if p.is_absolute():
        return p
    # Try relative to data/ directory
    resolved = _DATA_DIR / p
    if resolved.exists():
        return resolved
    # Fall back to CWD
    return Path(file_path).resolve()


def _read_csv(path: Path, max_rows: int) -> Dict[str, Any]:
    """Read a CSV file into structured data."""
    try:
        text = path.read_text(encoding="utf-8", errors="replace")
        reader = csv.reader(io.StringIO(text))
        all_rows: List[List[str]] = []
        for row in reader:
            all_rows.append([str(cell) for cell in row])

        headers = all_rows[0] if all_rows else []
        data_rows = all_rows[1:] if len(all_rows) > 1 else []
        truncated = len(data_rows) > max_rows
        data_rows = data_rows[:max_rows]

        return {
            "file": str(path),
            "format": "csv",
            "headers": headers,
            "rows": data_rows,
            "row_count": len(data_rows),
            "total_rows": len(all_rows) - 1,
            "truncated": truncated,
        }
    except Exception as e:
        logger.warning("csv_read_failed", path=str(path), error=str(e))
        return {"error": f"Failed to read CSV: {e}"}


def _read_xlsx(path: Path, sheet_name: Optional[str], max_rows: int) -> Dict[str, Any]:
    """Read an XLSX file into structured data."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        available_sheets = wb.sheetnames

        target_sheet = sheet_name or available_sheets[0]
        if target_sheet not in available_sheets:
            wb.close()
            return {
                "error": f"Sheet '{target_sheet}' not found. Available: {available_sheets}",
                "available_sheets": available_sheets,
            }

        ws = wb[target_sheet]
        all_rows: List[List[str]] = []
        for row in ws.iter_rows(values_only=True):
            all_rows.append([str(cell) if cell is not None else "" for cell in row])

        wb.close()

        headers = all_rows[0] if all_rows else []
        data_rows = all_rows[1:] if len(all_rows) > 1 else []
        truncated = len(data_rows) > max_rows
        data_rows = data_rows[:max_rows]

        return {
            "file": str(path),
            "format": "xlsx",
            "sheet": target_sheet,
            "available_sheets": available_sheets,
            "headers": headers,
            "rows": data_rows,
            "row_count": len(data_rows),
            "total_rows": len(all_rows) - 1,
            "truncated": truncated,
        }
    except Exception as e:
        logger.warning("xlsx_read_failed", path=str(path), error=str(e))
        return {"error": f"Failed to read XLSX: {e}"}
