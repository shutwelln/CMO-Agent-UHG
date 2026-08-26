"""Workspace management for multi-tenant brand support."""

from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Dict, List, Optional

import structlog

from ..db.database import Database
from ..db.models import Source, Workspace
from ..db.repositories import ConfigRepo, SourceRepo, WorkspaceRepo
from .brand_voice import BrandVoiceLoader

logger = structlog.get_logger()

# Default data directory where workspace folders live (e.g. data/uhg/)
DEFAULT_DATA_DIR = Path("./data")

# Max chars to extract per file to avoid blowing up context
_MAX_DOC_CHARS = 30_000
_MAX_TOTAL_DOCS_CHARS = 80_000

# Pattern for detecting workspace references in user messages
_WORKSPACE_PATTERN = re.compile(r"\bfor\s+([\w.-]+)\b", re.IGNORECASE)


class WorkspaceManager:
    """Manages workspaces, brand voices, and per-workspace configuration."""

    def __init__(
        self,
        db: Database,
        brand_voices_dir: Optional[Path] = None,
    ) -> None:
        self.db = db
        self._workspaces = WorkspaceRepo(db)
        self._config = ConfigRepo(db)
        self._sources = SourceRepo(db)
        self._voice_loader = BrandVoiceLoader(brand_voices_dir)

    async def get_workspace(self, workspace_id: str) -> Optional[Workspace]:
        return await self._workspaces.get(workspace_id)

    async def get_default_workspace(self) -> Optional[Workspace]:
        return await self._workspaces.get_default()

    async def list_workspaces(self) -> List[Workspace]:
        return await self._workspaces.list_all()

    async def create_workspace(
        self,
        workspace_id: str,
        name: str,
        workspace_type: str = "owned_brand",
        slack_channel: Optional[str] = None,
        brand_voice_path: Optional[str] = None,
        is_default: bool = False,
    ) -> Workspace:
        return await self._workspaces.create(
            workspace_id=workspace_id,
            name=name,
            workspace_type=workspace_type,
            slack_channel=slack_channel,
            brand_voice_path=brand_voice_path,
            is_default=is_default,
        )

    async def resolve_workspace(
        self, user_input: str, slack_channel: Optional[str] = None
    ) -> Workspace:
        """Resolve which workspace a user message targets.

        Resolution order:
        1. Explicit 'for <workspace>' pattern
        2. Workspace ID or name mentioned anywhere in the message text
        3. Slack channel match (if slack_channel provided)
        4. Default workspace fallback
        """
        input_lower = user_input.lower()

        # 1. Explicit "for <workspace>" pattern
        match = _WORKSPACE_PATTERN.search(user_input)
        if match:
            candidate = match.group(1).lower()
            ws = await self._workspaces.get(candidate)
            if ws:
                logger.debug("workspace_resolved", workspace_id=ws.id, method="explicit")
                return ws

            all_ws = await self._workspaces.list_all()
            for ws in all_ws:
                if ws.name.lower() == candidate or candidate in ws.name.lower():
                    logger.debug("workspace_resolved", workspace_id=ws.id, method="name_match")
                    return ws

        # 2. Scan for workspace ID or name anywhere in message text
        all_ws = await self._workspaces.list_all()
        for ws in all_ws:
            if ws.id in input_lower or ws.name.lower() in input_lower:
                logger.debug("workspace_resolved", workspace_id=ws.id, method="text_scan")
                return ws

        # 3. Match by Slack channel
        if slack_channel:
            for ws in all_ws:
                if ws.slack_channel and ws.slack_channel == slack_channel:
                    logger.debug("workspace_resolved", workspace_id=ws.id, method="channel_match")
                    return ws

        default = await self._workspaces.get_default()
        if default:
            logger.debug("workspace_resolved", workspace_id=default.id, method="default")
            return default

        raise ValueError("No default workspace configured")

    async def get_brand_voice(self, workspace_id: str) -> str:
        """Load the brand voice text for a workspace, with brand colors prepended."""
        ws = await self._workspaces.get(workspace_id)
        voice = ""
        if ws and ws.brand_voice_path:
            voice = self._voice_loader.load(ws.brand_voice_path)

        # Prepend brand colors so they survive truncation (brand_voice[:300])
        colors_raw = await self._config.get(workspace_id, "brand_colors")
        if colors_raw:
            try:
                colors = json.loads(colors_raw)
                lines = [f"- {role}: {val}" for role, val in colors.items()]
                voice = "BRAND COLORS:\n" + "\n".join(lines) + "\n\n" + voice
            except (json.JSONDecodeError, TypeError, AttributeError):
                pass

        return voice

    async def get_keywords(self, workspace_id: str) -> List[str]:
        """Get the keyword list for a workspace."""
        raw = await self._config.get(workspace_id, "keywords")
        if raw:
            return json.loads(raw)
        return []

    async def set_keywords(self, workspace_id: str, keywords: List[str]) -> None:
        await self._config.set(workspace_id, "keywords", json.dumps(keywords))

    async def get_config(self, workspace_id: str, key: str) -> Optional[str]:
        return await self._config.get(workspace_id, key)

    async def set_config(self, workspace_id: str, key: str, value: str) -> None:
        await self._config.set(workspace_id, key, value)

    async def get_reference_docs(self, workspace_id: str) -> str:
        """Load and parse reference documents from the workspace data directory.

        Scans data/{workspace_id}/ for PDFs, spreadsheets, and text files,
        extracts their content, and returns a combined text block suitable
        for injection into LLM prompts.

        Returns empty string if the directory doesn't exist or has no parseable files.
        """
        ws_dir = DEFAULT_DATA_DIR / workspace_id
        if not ws_dir.is_dir():
            return ""

        sections: List[str] = []
        total_chars = 0

        for file_path in sorted(ws_dir.iterdir()):
            if total_chars >= _MAX_TOTAL_DOCS_CHARS:
                sections.append(
                    f"[... additional documents in {ws_dir} truncated for context limit]"
                )
                break

            if file_path.is_dir() or file_path.name.startswith("."):
                continue

            text = self._extract_file_text(file_path)
            if not text:
                continue

            # Truncate individual docs
            if len(text) > _MAX_DOC_CHARS:
                text = text[:_MAX_DOC_CHARS] + "\n[... truncated]"

            section = f"--- {file_path.name} ---\n{text}"
            sections.append(section)
            total_chars += len(section)
            logger.debug(
                "reference_doc_loaded",
                workspace_id=workspace_id,
                file=file_path.name,
                chars=len(text),
            )

        if not sections:
            return ""

        combined = "\n\n".join(sections)
        logger.info(
            "reference_docs_loaded",
            workspace_id=workspace_id,
            file_count=len(sections),
            total_chars=len(combined),
        )
        return combined

    @staticmethod
    def _extract_file_text(file_path: Path) -> str:
        """Extract text content from a file based on its extension."""
        ext = file_path.suffix.lower()

        try:
            if ext == ".pdf":
                from pypdf import PdfReader

                reader = PdfReader(str(file_path))
                pages = []
                for i, page in enumerate(reader.pages):
                    text = page.extract_text() or ""
                    if text.strip():
                        pages.append(f"[Page {i + 1}] {text}")
                return "\n\n".join(pages)

            elif ext in (".xlsx", ".xls"):
                from openpyxl import load_workbook

                wb = load_workbook(str(file_path), read_only=True, data_only=True)
                sheets = []
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    rows = []
                    for i, row in enumerate(ws.iter_rows(values_only=True)):
                        if i >= 500:
                            rows.append("[... truncated at 500 rows]")
                            break
                        cells = [str(c) if c is not None else "" for c in row]
                        rows.append(" | ".join(cells))
                    if rows:
                        sheets.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows))
                wb.close()
                return "\n\n".join(sheets)

            elif ext == ".csv":
                return file_path.read_text(encoding="utf-8", errors="replace")[:_MAX_DOC_CHARS]

            elif ext in (".txt", ".md"):
                return file_path.read_text(encoding="utf-8", errors="replace")

            elif ext == ".docx":
                from docx import Document

                doc = Document(str(file_path))
                return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())

            else:
                return ""

        except Exception as e:
            logger.warning("reference_doc_parse_failed", file=str(file_path), error=str(e))
            return ""

    async def get_sources(
        self,
        workspace_id: str,
        source_type: Optional[str] = None,
    ) -> List[Source]:
        return await self._sources.list_by_workspace(workspace_id, source_type)
